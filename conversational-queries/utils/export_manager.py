import streamlit as st
import pandas as pd
import json
import time
from typing import Dict, Any, List, Optional

class ExportManager:
    """Gestionnaire amélioré pour les exports"""
    
    def __init__(self, results: Dict[str, Any], metadata: Dict[str, Any]):
        self.results = results
        self.metadata = metadata
        self.timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    import streamlit as st
import pandas as pd
import json
import time
from typing import Dict, Any, List, Optional
from io import BytesIO
from datetime import datetime

class ExportManager:
    """Gestionnaire amélioré pour les exports avec Excel professionnel"""
    
    def __init__(self, results: Dict[str, Any], metadata: Dict[str, Any]):
        self.results = results
        self.metadata = metadata
        self.timestamp = time.strftime("%Y%m%d_%H%M%S")
        self.export_date = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    def render_export_section(self):
        """Afficher la section d'export dans la sidebar"""
        if not self.results:
            return
        
        st.sidebar.markdown("---")
        st.sidebar.markdown("## 📥 Exports")
        
        # Exports Excel améliorés
        if self.results.get('all_suggestions') or self.results.get('enriched_keywords') or self.results.get('final_consolidated_data'):
            self._render_excel_exports()
        
        # Exports CSV traditionnels
        if self.results.get('all_suggestions'):
            self._render_suggestions_export()
        
        if self.results.get('enriched_keywords'):
            self._render_keywords_export()
        
        if self.results.get('final_consolidated_data'):
            self._render_questions_export()
        
        # Export complet (JSON)
        self._render_complete_export()
    
    def _render_excel_exports(self):
        """Exports Excel avec formatage professionnel"""
        st.sidebar.markdown("**📊 Exports Excel**")
        
        # Export Excel complet
        if st.sidebar.button("📈 Excel Complet", key="excel_complete", help="Toutes les données dans un fichier Excel multi-feuilles"):
            excel_data = self._create_complete_excel()
            if excel_data:
                st.sidebar.download_button(
                    label="📥 Télécharger Excel Complet",
                    data=excel_data,
                    file_name=f"analyse_complete_{self.timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_complete"
                )
        
        # Export Excel SEO optimisé
        if self.results.get('final_consolidated_data') and self.results.get('enriched_keywords'):
            if st.sidebar.button("🚀 Excel SEO", key="excel_seo", help="Questions + données de volume optimisées pour le SEO"):
                excel_data = self._create_seo_excel()
                if excel_data:
                    st.sidebar.download_button(
                        label="📥 Télécharger Excel SEO",
                        data=excel_data,
                        file_name=f"seo_questions_{self.timestamp}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel_seo"
                    )
        
        # Export Excel mots-clés
        if self.results.get('enriched_keywords'):
            if st.sidebar.button("🎯 Excel Mots-clés", key="excel_keywords", help="Analyse détaillée des mots-clés et volumes"):
                excel_data = self._create_keywords_excel()
                if excel_data:
                    st.sidebar.download_button(
                        label="📥 Télécharger Excel Mots-clés",
                        data=excel_data,
                        file_name=f"mots_cles_{self.timestamp}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel_keywords"
                    )
    
    def _create_complete_excel(self) -> Optional[BytesIO]:
        """Créer un fichier Excel complet avec toutes les données"""
        try:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                workbook = writer.book
                
                # Feuille 1: Résumé
                self._create_summary_sheet(writer)
                
                # Feuille 2: Suggestions Google
                if self.results.get('all_suggestions'):
                    self._create_suggestions_sheet(writer)
                
                # Feuille 3: Mots-clés avec volume
                if self.results.get('enriched_keywords'):
                    self._create_keywords_sheet(writer)
                
                # Feuille 4: Questions conversationnelles
                if self.results.get('final_consolidated_data'):
                    self._create_questions_sheet(writer)
                
                # Feuille 5: Analyse détaillée
                if self.results.get('enriched_keywords'):
                    self._create_analysis_sheet(writer)
            
            output.seek(0)
            return output
            
        except Exception as e:
            st.sidebar.error(f"Erreur création Excel complet: {str(e)}")
            return None
    
    def _create_summary_sheet(self, writer):
        """Créer la feuille de résumé"""
        summary_data = {
            'Métrique': [
                'Date d\'analyse',
                'Mots-clés analysés',
                'Suggestions collectées',
                'Mots-clés avec volume',
                'Questions générées',
                'Langue d\'analyse'
            ],
            'Valeur': [
                self.export_date,
                len(self.metadata.get('keywords', [])),
                len(self.results.get('all_suggestions', [])),
                len([k for k in self.results.get('enriched_keywords', []) if k.get('search_volume', 0) > 0]),
                len(self.results.get('final_consolidated_data', [])),
                self.metadata.get('language', 'fr').upper()
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Résumé', index=False)
        
        # Formatage
        worksheet = writer.sheets['Résumé']
        self._apply_excel_formatting(worksheet, df_summary)
    
    def _create_suggestions_sheet(self, writer):
        """Créer la feuille des suggestions"""
        df = pd.DataFrame(self.results['all_suggestions'])
        df.to_excel(writer, sheet_name='Suggestions', index=False)
        
        worksheet = writer.sheets['Suggestions']
        self._apply_excel_formatting(worksheet, df)
    
    def _create_keywords_sheet(self, writer):
        """Créer la feuille des mots-clés"""
        df = pd.DataFrame(self.results['enriched_keywords'])
        
        # Sélectionner et renommer les colonnes importantes
        columns_mapping = {
            'keyword': 'Mot-clé',
            'search_volume': 'Volume/mois',
            'cpc': 'CPC',
            'competition_level': 'Niveau_concurrence',
            'origine': 'Origine'
        }
        
        available_cols = [col for col in columns_mapping.keys() if col in df.columns]
        df_export = df[available_cols].copy()
        df_export = df_export.rename(columns=columns_mapping)
        
        # Trier par volume décroissant
        if 'Volume/mois' in df_export.columns:
            df_export = df_export.sort_values('Volume/mois', ascending=False)
        
        df_export.to_excel(writer, sheet_name='Mots-clés', index=False)
        
        worksheet = writer.sheets['Mots-clés']
        self._apply_excel_formatting(worksheet, df_export)
    
    def _create_questions_sheet(self, writer):
        """Créer la feuille des questions"""
        df = pd.DataFrame(self.results['final_consolidated_data'])
        df.to_excel(writer, sheet_name='Questions', index=False)
        
        worksheet = writer.sheets['Questions']
        self._apply_excel_formatting(worksheet, df)
    
    def _create_analysis_sheet(self, writer):
        """Créer la feuille d'analyse détaillée"""
        enriched_keywords = self.results.get('enriched_keywords', [])
        
        # Statistiques par origine
        origins = {}
        for kw in enriched_keywords:
            origin = kw.get('origine', 'Inconnue')
            if origin not in origins:
                origins[origin] = []
            origins[origin].append(kw)
        
        analysis_data = []
        for origin, keywords in origins.items():
            volumes = [k.get('search_volume', 0) for k in keywords if k.get('search_volume', 0) > 0]
            analysis_data.append({
                'Origine': origin,
                'Nombre_mots_cles': len(keywords),
                'Avec_volume': len(volumes),
                'Volume_total': sum(volumes),
                'Volume_moyen': sum(volumes) / len(volumes) if volumes else 0,
                'Volume_max': max(volumes) if volumes else 0
            })
        
        df_analysis = pd.DataFrame(analysis_data)
        df_analysis.to_excel(writer, sheet_name='Analyse', index=False)
        
        worksheet = writer.sheets['Analyse']
        self._apply_excel_formatting(worksheet, df_analysis)
    
    def _create_seo_excel(self) -> Optional[BytesIO]:
        """Créer un fichier Excel optimisé pour le SEO"""
        try:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # Feuille principale: Questions SEO
                seo_df = self._create_seo_optimized_export()
                if seo_df is not None:
                    seo_df.to_excel(writer, sheet_name='Questions_SEO', index=False)
                    worksheet = writer.sheets['Questions_SEO']
                    self._apply_excel_formatting(worksheet, seo_df)
                
                # Feuille: Top mots-clés
                if self.results.get('enriched_keywords'):
                    top_keywords = self._get_top_keywords_for_seo()
                    if not top_keywords.empty:
                        top_keywords.to_excel(writer, sheet_name='Top_Mots_clés', index=False)
                        worksheet = writer.sheets['Top_Mots_clés']
                        self._apply_excel_formatting(worksheet, top_keywords)
            
            output.seek(0)
            return output
            
        except Exception as e:
            st.sidebar.error(f"Erreur création Excel SEO: {str(e)}")
            return None
    
    def _create_keywords_excel(self) -> Optional[BytesIO]:
        """Créer un fichier Excel spécialisé mots-clés"""
        try:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # Feuille: Tous les mots-clés
                if self.results.get('enriched_keywords'):
                    df = pd.DataFrame(self.results['enriched_keywords'])
                    df.to_excel(writer, sheet_name='Tous_les_mots_clés', index=False)
                    worksheet = writer.sheets['Tous_les_mots_clés']
                    self._apply_excel_formatting(worksheet, df)
                
                # Feuille: Mots-clés avec volume
                keywords_with_volume = [k for k in self.results.get('enriched_keywords', []) if k.get('search_volume', 0) > 0]
                if keywords_with_volume:
                    df_volume = pd.DataFrame(keywords_with_volume)
                    df_volume = df_volume.sort_values('search_volume', ascending=False)
                    df_volume.to_excel(writer, sheet_name='Avec_volume', index=False)
                    worksheet = writer.sheets['Avec_volume']
                    self._apply_excel_formatting(worksheet, df_volume)
                
                # Feuille: Statistiques
                if self.results.get('enriched_keywords'):
                    stats_df = self._create_keywords_statistics()
                    if not stats_df.empty:
                        stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
                        worksheet = writer.sheets['Statistiques']
                        self._apply_excel_formatting(worksheet, stats_df)
            
            output.seek(0)
            return output
            
        except Exception as e:
            st.sidebar.error(f"Erreur création Excel mots-clés: {str(e)}")
            return None
    
    def _get_top_keywords_for_seo(self) -> pd.DataFrame:
        """Obtenir les top mots-clés pour l'export SEO"""
        enriched_keywords = self.results.get('enriched_keywords', [])
        keywords_with_volume = [k for k in enriched_keywords if k.get('search_volume', 0) > 0]
        
        if not keywords_with_volume:
            return pd.DataFrame()
        
        # Trier par volume et prendre les 50 premiers
        sorted_keywords = sorted(keywords_with_volume, key=lambda x: x.get('search_volume', 0), reverse=True)[:50]
        
        df = pd.DataFrame(sorted_keywords)
        columns_mapping = {
            'keyword': 'Mot_clé',
            'search_volume': 'Volume_mensuel',
            'cpc': 'CPC_estime',
            'competition_level': 'Concurrence',
            'origine': 'Origine'
        }
        
        available_cols = [col for col in columns_mapping.keys() if col in df.columns]
        df_export = df[available_cols].copy()
        df_export = df_export.rename(columns=columns_mapping)
        
        return df_export
    
    def _create_keywords_statistics(self) -> pd.DataFrame:
        """Créer des statistiques sur les mots-clés"""
        enriched_keywords = self.results.get('enriched_keywords', [])
        
        if not enriched_keywords:
            return pd.DataFrame()
        
        # Statistiques générales
        all_volumes = [k.get('search_volume', 0) for k in enriched_keywords]
        volumes_with_data = [v for v in all_volumes if v > 0]
        
        stats = {
            'Métrique': [
                'Total mots-clés',
                'Avec volume de recherche',
                'Sans volume de recherche',
                'Volume total (mensuel)',
                'Volume moyen',
                'Volume médian',
                'Volume maximum'
            ],
            'Valeur': [
                len(enriched_keywords),
                len(volumes_with_data),
                len(enriched_keywords) - len(volumes_with_data),
                sum(volumes_with_data) if volumes_with_data else 0,
                sum(volumes_with_data) / len(volumes_with_data) if volumes_with_data else 0,
                sorted(volumes_with_data)[len(volumes_with_data)//2] if volumes_with_data else 0,
                max(volumes_with_data) if volumes_with_data else 0
            ]
        }
        
        return pd.DataFrame(stats)
    
    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame):
        """Appliquer un formatage professionnel à une feuille Excel"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Formatage de l'en-tête
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, cell in enumerate(worksheet[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Maximum 50 caractères
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Formatage des nombres
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
        
        except ImportError:
            # Si openpyxl n'est pas disponible, passer le formatage
            pass
    
    def _render_suggestions_export(self):
        """Export des suggestions"""
        suggestions_df = pd.DataFrame(self.results['all_suggestions'])
        
        # CSV
        csv_suggestions = suggestions_df.to_csv(index=False)
        st.sidebar.download_button(
            label="📝 Suggestions (CSV)",
            data=csv_suggestions,
            file_name=f"suggestions_{self.timestamp}.csv",
            mime="text/csv",
            help="Toutes les suggestions Google collectées"
        )
        
        # TXT (liste simple)
        txt_suggestions = "\n".join(suggestions_df['Suggestion Google'].tolist())
        st.sidebar.download_button(
            label="📄 Suggestions (TXT)",
            data=txt_suggestions,
            file_name=f"suggestions_{self.timestamp}.txt",
            mime="text/plain",
            help="Liste simple des suggestions"
        )
    
    def _render_keywords_export(self):
        """Export des mots-clés avec volume"""
        enriched_keywords = self.results['enriched_keywords']
        keywords_df = pd.DataFrame(enriched_keywords)
        
        # Préparer les colonnes pour l'export
        export_cols = ['keyword', 'search_volume', 'cpc', 'competition', 'competition_level', 'source', 'origine']
        available_cols = [col for col in export_cols if col in keywords_df.columns]
        export_df = keywords_df[available_cols].copy()
        
        # Renommer les colonnes
        column_mapping = {
            'keyword': 'Mot-clé',
            'search_volume': 'Volume/mois',
            'cpc': 'CPC',
            'competition': 'Concurrence',
            'competition_level': 'Niveau_Concurrence',
            'source': 'Source',
            'origine': 'Origine'
        }
        export_df = export_df.rename(columns=column_mapping)
        
        # CSV des mots-clés enrichis
        csv_keywords = export_df.to_csv(index=False)
        st.sidebar.download_button(
            label="📊 Mots-clés + Volumes (CSV)",
            data=csv_keywords,
            file_name=f"keywords_volumes_{self.timestamp}.csv",
            mime="text/csv",
            help="Mots-clés avec volumes de recherche et données DataForSEO"
        )
        
        # Export des mots-clés avec volume uniquement
        keywords_with_volume = export_df[export_df['Volume/mois'] > 0].copy()
        if not keywords_with_volume.empty:
            csv_volume_only = keywords_with_volume.to_csv(index=False)
            st.sidebar.download_button(
                label="🎯 Mots-clés avec volume (CSV)",
                data=csv_volume_only,
                file_name=f"keywords_with_volume_{self.timestamp}.csv",
                mime="text/csv",
                help="Uniquement les mots-clés avec volume de recherche"
            )
    
    def _render_questions_export(self):
        """Export des questions conversationnelles"""
        questions_df = pd.DataFrame(self.results['final_consolidated_data'])
        
        # CSV des questions
        csv_questions = questions_df.to_csv(index=False)
        st.sidebar.download_button(
            label="✨ Questions conversationnelles (CSV)",
            data=csv_questions,
            file_name=f"questions_{self.timestamp}.csv",
            mime="text/csv",
            help="Questions conversationnelles générées"
        )
        
        # Export optimisé pour SEO (questions + volumes)
        if self.results.get('enriched_keywords'):
            seo_export = self._create_seo_optimized_export(questions_df)
            if seo_export is not None:
                csv_seo = seo_export.to_csv(index=False)
                st.sidebar.download_button(
                    label="🚀 Export SEO optimisé (CSV)",
                    data=csv_seo,
                    file_name=f"seo_questions_{self.timestamp}.csv",
                    mime="text/csv",
                    help="Questions avec données de volume pour optimisation SEO"
                )
    
    def _create_seo_optimized_export(self, questions_df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Créer un export optimisé pour le SEO"""
        try:
            enriched_df = pd.DataFrame(self.results['enriched_keywords'])
            if enriched_df.empty or 'keyword' not in enriched_df.columns:
                return None
            
            # Merger avec les données de volume
            merged_df = questions_df.merge(
                enriched_df[['keyword', 'search_volume', 'cpc', 'competition_level', 'origine']],
                left_on='Suggestion Google',
                right_on='keyword',
                how='left'
            )
            
            # Sélectionner et renommer les colonnes pour le SEO
            seo_cols = {
                'Question Conversationnelle': 'Question_SEO',
                'Suggestion Google': 'Mot_cle_cible',
                'Thème': 'Theme',
                'Intention': 'Intention_recherche',
                'Score_Importance': 'Score_importance',
                'search_volume': 'Volume_mensuel',
                'cpc': 'CPC_estime',
                'competition_level': 'Niveau_concurrence',
                'origine': 'Source_mot_cle'
            }
            
            available_seo_cols = {k: v for k, v in seo_cols.items() if k in merged_df.columns}
            seo_export = merged_df[list(available_seo_cols.keys())].copy()
            seo_export = seo_export.rename(columns=available_seo_cols)
            
            # Formater les données
            if 'Volume_mensuel' in seo_export.columns:
                seo_export['Volume_mensuel'] = seo_export['Volume_mensuel'].fillna(0).astype(int)
            if 'CPC_estime' in seo_export.columns:
                seo_export['CPC_estime'] = seo_export['CPC_estime'].fillna(0).round(2)
            
            # Trier par volume décroissant puis par score d'importance
            sort_cols = []
            if 'Volume_mensuel' in seo_export.columns:
                sort_cols.append('Volume_mensuel')
            if 'Score_importance' in seo_export.columns:
                sort_cols.append('Score_importance')
            
            if sort_cols:
                seo_export = seo_export.sort_values(sort_cols, ascending=False)
            
            return seo_export
            
        except Exception as e:
            st.sidebar.error(f"Erreur export SEO: {str(e)}")
            return None
    
    def _render_complete_export(self):
        """Export complet au format JSON"""
        complete_data = {
            'metadata': self.metadata,
            'results': {
                'suggestions': self.results.get('all_suggestions', []),
                'enriched_keywords': self.results.get('enriched_keywords', []),
                'questions': self.results.get('final_consolidated_data', []),
                'themes_analysis': self.results.get('themes_analysis', {}),
                'selected_themes': self.results.get('selected_themes_by_keyword', {})
            },
            'export_timestamp': self.timestamp
        }
        
        json_data = json.dumps(complete_data, ensure_ascii=False, indent=2)
        st.sidebar.download_button(
            label="📦 Export complet (JSON)",
            data=json_data,
            file_name=f"analysis_complete_{self.timestamp}.json",
            mime="application/json",
            help="Toutes les données de l'analyse au format JSON"
        )
